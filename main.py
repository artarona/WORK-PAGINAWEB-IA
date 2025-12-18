# pip install -r requirements.txt

import sys
import os
import json
from flask import Flask, request, jsonify
from flask_cors import CORS
from openpyxl import Workbook, load_workbook
from datetime import datetime

# Importar lógica de la base de datos de la IA
from logic.database import query_properties, initialize_databases, get_historial_canal, log_conversation
from logic.filters import detect_filters
from logic.gemini_client import call_gemini_with_rotation, build_prompt
import time

app = Flask(__name__)
CORS(app, resources={r"/api/*": {"origins": ["null", "http://dantepropiedades.com.ar", "http://www.dantepropiedades.com.ar", "https://dantepropiedades.com.ar", "https://www.dantepropiedades.com.ar", "http://dantepropiedades.com", "https://danterealestate-github-io.onrender.com", "https://danterealestate.github.io"]}})

EXCEL_FILE = 'contactos_dante_propiedades.xlsx'

def safe_print(message):
    safe_message = message.encode('ascii', 'ignore').decode('ascii')
    print(safe_message)

def init_excel():
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Contactos"
        ws['A1'] = 'Fecha/Hora'
        ws['B1'] = 'Nombre'
        ws['C1'] = 'Firma'
        ws['D1'] = 'Teléfono'
        ws['E1'] = 'Propiedad'
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        wb.save(EXCEL_FILE)
        safe_print(f"SUCCESS: Archivo {EXCEL_FILE} creado exitosamente")
    else:
        safe_print(f"INFO: Archivo {EXCEL_FILE} encontrado")

def serve_static_file(filename):
    try:
        if filename.startswith('api/') or filename in ['contactos_dante_propiedades.xlsx', 'propiedades.json']:
            safe_print(f"Archivo {filename} no se sirve como estático - será manejado por endpoint específico")
            return jsonify({"error": f"Archivo {filename} no encontrado"}), 404
        
        file_path = os.path.join(os.getcwd(), filename)
        safe_print(f"Buscando archivo: {file_path}")
        
        if os.path.exists(file_path):
            try:
                with open(file_path, 'rb') as f:
                    content = f.read()
                
                # Determinar el tipo de contenido
                if filename.lower().endswith(('.png', '.jpg', '.jpeg', '.gif', '.ico', '.svg')):
                    from flask import Response
                    return Response(content, mimetype=f'image/{filename.split(".")[-1].lower()}')
                elif filename.lower().endswith('.css'):
                    from flask import Response
                    return Response(content, mimetype='text/css')
                elif filename.lower().endswith('.js'):
                    from flask import Response
                    return Response(content, mimetype='application/javascript')
                elif filename.lower().endswith('.html'):
                    from flask import Response
                    return Response(content, mimetype='text/html')
                elif filename.lower().endswith('.json'):
                    from flask import Response
                    return Response(content, mimetype='application/json')
                else:
                    from flask import Response
                    return Response(content, mimetype='application/octet-stream')
                    
            except Exception as e:
                safe_print(f"Error leyendo {filename}: {str(e)}")
                return jsonify({"error": f"Error al leer archivo {filename}"}), 500
        else:
            safe_print(f"Archivo NO encontrado: {file_path}")
            return jsonify({"error": f"Archivo {filename} no encontrado"}), 404
            
    except Exception as e:
        safe_print(f"Error sirviendo {filename}: {str(e)}")
        return jsonify({"error": f"Error al servir {filename}"}), 500

@app.route('/')
def home():
    return serve_static_file('index.html')

@app.route('/api/guardar_contacto', methods=['POST'])
def guardar_contacto():
    try:
        data = request.get_json()
        
        if not data or 'nombre' not in data:
            return jsonify({"error": "Datos incompletos"}), 400
        
        init_excel()
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        next_row = ws.max_row + 1
        now = datetime.now()
        fecha_hora = now.strftime("%Y-%m-%d %H:%M:%S")
        
        ws[f'A{next_row}'] = fecha_hora
        ws[f'B{next_row}'] = data.get('nombre', '')
        ws[f'C{next_row}'] = data.get('email', '')
        ws[f'D{next_row}'] = data.get('telefono', '')
        ws[f'E{next_row}'] = data.get('propiedad_interes', '')
        
        wb.save(EXCEL_FILE)
        safe_print(f"Contacto guardado: {data.get('nombre')} - {data.get('email')}")
        
        return jsonify({
            "message": "Contacto guardado exitosamente",
            "fecha_hora": fecha_hora
        }), 200
        
    except Exception as e:
        safe_print(f"Error guardando contacto: {str(e)}")
        return jsonify({"error": f"Error al guardar contacto: {str(e)}"}), 500

@app.route('/api/properties', methods=['GET'])
def get_all_properties():
    try:
        properties = query_properties({})
        return jsonify({
            "total": len(properties),
            "properties": properties
        }), 200
    except Exception as e:
        safe_print(f"Error obteniendo propiedades: {str(e)}")
        return jsonify({"error": f"Error al obtener propiedades"}), 500

def get_filter_options():
    try:
        properties = query_properties({})
        barrios = []
        tipos = []
        
        for prop in properties:
            if not isinstance(prop, dict):
                safe_print(f"ADVERTENCIA: Elemento no es diccionario: {prop}")
                continue
            
            if prop.get('barrio') and prop['barrio'] not in barrios:
                barrios.append(prop['barrio'])
            
            if prop.get('tipo') and prop['tipo'] not in tipos:
                tipos.append(prop['tipo'])
        
        return {
            "barrios": sorted(barrios),
            "tipos": sorted(tipos),
            "total": len(properties)
        }
        
    except Exception as e:
        safe_print(f"Error en get_filter_options: {str(e)}")
        return {"error": f"Error obteniendo opciones de filtros: {str(e)}"}

@app.route('/api/properties/filter-options', methods=['GET'])
def get_filter_options_endpoint():
    try:
        options = get_filter_options()
        if "error" in options:
            return jsonify(options), 500
        
        return jsonify(options), 200
        
    except Exception as e:
        safe_print(f"Error en endpoint filter-options: {str(e)}")
        return jsonify({"error": f"Error en servidor: {str(e)}"}), 500

@app.route('/api/properties/search', methods=['GET'])
def search_properties_endpoint():
    try:
        # Mapeo de los parámetros del frontend a las claves de filtro de la BD
        filters = {
            'operacion': request.args.get('ope'),
            'tipo': request.args.get('tipo'),
            'barrio': request.args.get('loc')
        }
        
        # Añadir filtros numéricos solo si tienen valor
        precio_max = request.args.get('precio_max')
        if precio_max:
            try:
                filters['max_price'] = float(precio_max)
            except ValueError:
                pass # Ignorar si no es un número válido

        ambientes = request.args.get('ambientes')
        if ambientes:
            try:
                filters['min_rooms'] = int(ambientes)
            except ValueError:
                pass # Ignorar si no es un número válido

        # Limpiar filtros nulos
        active_filters = {k: v for k, v in filters.items() if v is not None}
        
        safe_print(f"--- Nueva Búsqueda ---")
        safe_print(f"Filtros activos: {active_filters}")
        
        results = query_properties(active_filters)
        
        safe_print(f"Propiedades encontradas: {len(results)}")
        
        return jsonify({
            "total": len(results),
            "properties": results,
            "filters": active_filters
        }), 200
        
    except Exception as e:
        safe_print(f"Error en endpoint search: {str(e)}")
        return jsonify({"error": f"Error en servidor: {str(e)}"}), 500

class Metrics:
    def __init__(self):
        self.requests_count = 0
        self.successful_requests = 0
        self.failed_requests = 0
        self.gemini_calls = 0
        self.search_queries = 0
        self.start_time = time.time()
    
    def increment_requests(self): self.requests_count += 1
    def increment_success(self): self.successful_requests += 1
    def increment_failures(self): self.failed_requests += 1
    def increment_gemini_calls(self): self.gemini_calls += 1
    def increment_searches(self): self.search_queries += 1
    def get_uptime(self): return time.time() - self.start_time

metrics = Metrics()

@app.route('/api/status')
def status():
    return jsonify({
        "status": "activo",
        "uptime_seconds": metrics.get_uptime(),
        "total_requests": metrics.requests_count,
        "gemini_calls": metrics.gemini_calls,
        "search_queries": metrics.search_queries
    })

@app.route('/api/chat', methods=['POST'])
def chat():
    start_time = time.time()
    try:
        data = request.get_json()
        if not data or not data.get('message'):
            return jsonify({"error": "El mensaje no puede estar vacío"}), 400

        user_text = data.get('message', '').strip()
        channel = data.get('channel', 'web').strip()
        filters_from_frontend = data.get('filters', {})
        contexto_anterior = data.get('contexto_anterior')

        text_lower = user_text.lower()
        filters = filters_from_frontend.copy()
        detected_filters = detect_filters(text_lower)
        filters.update(detected_filters)

        safe_print(f"--- Nueva Consulta de Chat ---")
        safe_print(f"Usuario: '{user_text}'")
        safe_print(f"Filtros combinados: {filters}")

        results = None
        search_performed = False
        
        if filters:
            search_performed = True
            results = query_properties(filters)
            safe_print(f"Resultados de la búsqueda: {len(results) if results else 0} propiedades")

        historial = get_historial_canal(channel)
        
        prompt = build_prompt(user_text, results, filters, channel, historial)
        answer = call_gemini_with_rotation(prompt)
        
        response_time = time.time() - start_time
        log_conversation(user_text, answer, channel, response_time, search_performed, len(results) if results else 0)
        
        response_data = {
            "response": answer,
            "results_count": len(results) if results is not None else None,
            "search_performed": search_performed,
            "propiedades": results
        }
        
        return jsonify(response_data), 200
    
    except Exception as e:
        safe_print(f"❌ ERROR en endpoint /api/chat: {type(e).__name__}: {e}")
        return jsonify({"error": "Ocurrió un error procesando tu consulta."}), 500

@app.route('/api/debug-images')
def debug_images():
    """Endpoint para verificar qué imágenes están disponibles"""
    import os
    try:
        if os.path.exists("imgs"):
            image_files = os.listdir("imgs")
            return jsonify({
                "message": "Carpeta imgs encontrada",
                "path_absoluto": os.path.abspath("imgs"),
                "total_images": len(image_files),
                "images": sorted(image_files)[:20]  # Primeras 20 imágenes ordenadas
            })
        else:
            return jsonify({"error": "Carpeta 'imgs' no encontrada en el servidor"})
    except Exception as e:
        return jsonify({"error": f"Error al leer carpeta: {str(e)}"})


@app.route('/api/status', methods=['GET'])
def api_status():
    """Endpoint para verificar el estado del servidor"""
    return jsonify({
        "status": "online",
        "message": "Servidor funcionando correctamente",
        "timestamp": datetime.now().isoformat()
    }), 200

# **RUTA GENÉRICA AL FINAL** - Debe ser la última
@app.route('/<path:filename>')
def serve_any_file(filename):
    return serve_static_file(filename)

if __name__ == '__main__':
    safe_print("Iniciando servidor Flask...")
    initialize_databases() # Asegura que la BD SQLite esté lista
    init_excel()
    safe_print("Servidor listo para recibir solicitudes")
    
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)